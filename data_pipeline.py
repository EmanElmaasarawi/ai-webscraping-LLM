"""Data pipeline for processing and exporting scraped data to Excel."""
import logging
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

logger = logging.getLogger(__name__)


class DataPipeline:
    """Pipeline for processing and exporting product data."""

    def __init__(self, output_folder: str = 'output', filename: str = 'scraped_products.xlsx'):
        """Initialize data pipeline."""
        self.output_folder = output_folder
        self.filename = filename
        self.open_spider(None)

    def open_spider(self, spider):
        self.data = []
        self.stats = {
            'total_items': 0,
            'valid_items': 0,
            'invalid_items': 0,
            'websites_scraped': set()
        }
    
    def process_item(self, item, spider):
        website = item.get("website", "unknown")

        if not self._validate_item(item):
            self.stats['invalid_items'] += 1
            return item

        item = self._normalize_item(item)

        self.data.append(item)

        self.stats['valid_items'] += 1
        self.stats['total_items'] += 1
        self.stats['websites_scraped'].add(website)

        return item
    def process_item2(self, item: Dict[str, Any], website: str) -> bool:
        """
        Process a scraped item and add to pipeline.
        
        Args:
            item: Dictionary with product data
            website: Name of the website
            
        Returns:
            True if item is valid, False otherwise
        """
        try:
            # Validate required fields
            if not self._validate_item(item):
                self.stats['invalid_items'] += 1
                logger.warning(f"Invalid item from {website}: {item}")
                return False
            
            # Add metadata
            item['website'] = website
            item['scraped_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Normalize data
            item = self._normalize_item(item)
            
            self.data.append(item)
            self.stats['valid_items'] += 1
            self.stats['total_items'] += 1
            self.stats['websites_scraped'].add(website)
            
            return True
            
        except Exception as e:
            logger.error(f"Error processing item: {e}")
            self.stats['invalid_items'] += 1
            return False

    def _validate_item(self, item: Dict[str, Any]) -> bool:
        """Validate item has required fields and data."""
        required_fields = ['name', 'price']
        
        for field in required_fields:
            if field not in item or not item[field]:
                return False
        
        # Price should be numeric
        try:
            float(item['price'])
        except (ValueError, TypeError):
            return False
        
        return True

    def _normalize_item(self, item: Dict[str, Any]) -> Dict[str, Any]:
        """Normalize item data."""
        # Clean up price
        if isinstance(item['price'], str):
            # Remove currency symbols and extra spaces
            item['price'] = float(''.join(c for c in item['price'] if c.isdigit() or c == '.'))
        
        # Clean up discount
        if item.get('discount') and isinstance(item['discount'], str):
            item['discount'] = item['discount'].strip()
        
        # Clean up unit
        if item.get('unit') and isinstance(item['unit'], str):
            item['unit'] = item['unit'].strip()
        
        # Convert quantity to int
        try:
            item['quantity'] = int(float(str(item.get('quantity', 0))))
        except (ValueError, TypeError):
            item['quantity'] = 0
        
        # Ensure all expected fields exist
        for field in ['name', 'price', 'discount', 'unit', 'quantity', 'website', 'scraped_at']:
            if field not in item:
                item[field] = 'N/A' if field in ['discount', 'unit'] else 0
        
        return item
    
    def close_spider(self, spider):
        self.export_to_excel()
    
    def export_to_excel(self) -> str:
        """
        Export all collected data to Excel file.
        
        Returns:
            Path to the created Excel file
        """
        try:
            os.makedirs(self.output_folder, exist_ok=True)
            filepath = os.path.join(self.output_folder, self.filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Products'
            
            # Set up headers
            headers = ['Name', 'Price', 'Discount', 'Unit', 'Quantity', 'Website', 'Scraped At']
            ws.append(headers)
            
            # Style header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Add data rows
            for item in self.data:
                ws.append([
                    item.get('name', 'N/A'),
                    item.get('price', 0),
                    item.get('discount', 'N/A'),
                    item.get('unit', 'N/A'),
                    item.get('quantity', 0),
                    item.get('website', 'N/A'),
                    item.get('scraped_at', 'N/A')
                ])
            
            # Style data rows
            for row in ws.iter_rows(min_row=2, max_row=len(self.data) + 1, min_col=1, max_col=7):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 35  # Name
            ws.column_dimensions['B'].width = 12  # Price
            ws.column_dimensions['C'].width = 15  # Discount
            ws.column_dimensions['D'].width = 12  # Unit
            ws.column_dimensions['E'].width = 12  # Quantity
            ws.column_dimensions['F'].width = 15  # Website
            ws.column_dimensions['G'].width = 20  # Scraped At
            
            # Add summary sheet
            summary_ws = wb.create_sheet('Summary')
            summary_ws.append(['Scraping Summary', ''])
            summary_ws.append(['Total Items Scraped', self.stats['total_items']])
            summary_ws.append(['Valid Items', self.stats['valid_items']])
            summary_ws.append(['Invalid Items', self.stats['invalid_items']])
            summary_ws.append(['Websites Scraped', ', '.join(self.stats['websites_scraped'])])
            summary_ws.append(['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 35
            
            wb.save(filepath)
            logger.info(f"Data exported to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise

    def get_statistics(self) -> Dict[str, Any]:
        """Get scraping statistics."""
        return {
            'total_items': self.stats['total_items'],
            'valid_items': self.stats['valid_items'],
            'invalid_items': self.stats['invalid_items'],
            'websites': list(self.stats['websites_scraped']),
            'data_count': len(self.data)
        }

    def clear_data(self) -> None:
        """Clear all collected data."""
        self.data = []
        self.stats = {
            'total_items': 0,
            'valid_items': 0,
            'invalid_items': 0,
            'websites_scraped': set()
        }
