"""Test script for the web scraper components - works without API key."""
import logging
import sys
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def test_config():
    """Test configuration loading."""
    logger.info("Testing configuration...")
    try:
        from config import WEBSITES, OPENAI_API_KEY, LLM_MODEL
        logger.info(f"✓ Websites configured: {list(WEBSITES.keys())}")
        logger.info(f"✓ LLM Model: {LLM_MODEL}")
        return True
    except Exception as e:
        logger.error(f"✗ Configuration error: {e}")
        return False


def test_data_pipeline():
    """Test data pipeline with sample data."""
    logger.info("\nTesting data pipeline...")
    try:
        from data_pipeline import DataPipeline
        
        # Create pipeline
        pipeline = DataPipeline(output_folder='output', filename='test_products.xlsx')
        
        # Sample products from hyperone/carrefour
        sample_data = [
            {
                'name': 'Fresh Tomatoes',
                'price': '2.99',
                'discount': '15%',
                'unit': 'kg',
                'quantity': '150'
            },
            {
                'name': 'Organic Bananas',
                'price': 1.49,
                'discount': None,
                'unit': 'bundle',
                'quantity': 200
            },
            {
                'name': 'Whole Milk',
                'price': '3.50',
                'discount': '5%',
                'unit': '1L',
                'quantity': '80'
            },
            {
                'name': 'Cheddar Cheese',
                'price': '5.99',
                'discount': '10%',
                'unit': '250g',
                'quantity': '45'
            },
            {
                'name': 'Brown Bread',
                'price': '2.25',
                'discount': None,
                'unit': '800g',
                'quantity': '120'
            },
        ]
        
        # Process items from hyperone
        logger.info("Processing hyperone products...")
        for item in sample_data[:3]:
            if pipeline.process_item(item, 'hyperone'):
                logger.info(f"  ✓ {item['name']}")
            else:
                logger.warning(f"  ✗ Failed to process: {item['name']}")
        
        # Process items from carrefour
        logger.info("Processing carrefour products...")
        for item in sample_data[3:]:
            if pipeline.process_item(item, 'carrefour'):
                logger.info(f"  ✓ {item['name']}")
            else:
                logger.warning(f"  ✗ Failed to process: {item['name']}")
        
        # Export to Excel
        logger.info("\nExporting to Excel...")
        filepath = pipeline.export_to_excel()
        logger.info(f"✓ Excel file created: {filepath}")
        
        # Show statistics
        stats = pipeline.get_statistics()
        logger.info(f"\n✓ Pipeline Statistics:")
        logger.info(f"  - Total items: {stats['total_items']}")
        logger.info(f"  - Valid items: {stats['valid_items']}")
        logger.info(f"  - Invalid items: {stats['invalid_items']}")
        logger.info(f"  - Websites: {stats['websites']}")
        
        return True
        
    except Exception as e:
        logger.error(f"✗ Pipeline error: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_selector_cache():
    """Test selector cache."""
    logger.info("\nTesting selector cache...")
    try:
        from llm_agent import SelectorCache
        
        cache = SelectorCache()
        
        # Test set and get
        test_selectors = {
            'product_item': 'div.product',
            'name_selector': 'h2.name',
            'price_selector': 'span.price'
        }
        
        cache.set_selectors('test-site', test_selectors)
        logger.info("✓ Selectors cached for test-site")
        
        retrieved = cache.get_selectors('test-site')
        if retrieved == test_selectors:
            logger.info("✓ Selectors retrieved correctly")
        else:
            logger.error("✗ Selectors mismatch")
            return False
        
        return True
        
    except Exception as e:
        logger.error(f"✗ Cache error: {e}")
        return False


def test_beautifulsoup():
    """Test HTML parsing capability."""
    logger.info("\nTesting HTML parsing...")
    try:
        from bs4 import BeautifulSoup
        
        sample_html = """
        <div class="product-grid">
            <div class="product-item">
                <h2 class="product-name">Fresh Apples</h2>
                <span class="price">2.99 USD</span>
                <span class="discount">15% off</span>
                <span class="unit">1 kg</span>
                <span class="quantity">In stock: 50</span>
            </div>
        </div>
        """
        
        soup = BeautifulSoup(sample_html, 'html.parser')
        
        # Test selectors
        name = soup.select_one('.product-name')
        price = soup.select_one('.price')
        discount = soup.select_one('.discount')
        
        if name and 'Apples' in name.text:
            logger.info(f"✓ Found product: {name.text.strip()}")
        else:
            logger.error("✗ Could not parse product name")
            return False
        
        if price and '2.99' in price.text:
            logger.info(f"✓ Found price: {price.text.strip()}")
        else:
            logger.error("✗ Could not parse price")
            return False
        
        return True
        
    except Exception as e:
        logger.error(f"✗ HTML parsing error: {e}")
        return False


def test_excel_creation():
    """Test Excel file creation directly."""
    logger.info("\nTesting Excel creation...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test'
        
        # Add headers
        ws['A1'] = 'Product'
        ws['B1'] = 'Price'
        ws['C1'] = 'Quantity'
        
        # Style
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws[cell].font = Font(bold=True, color='FFFFFF')
        
        # Add data
        ws['A2'] = 'Test Product'
        ws['B2'] = 9.99
        ws['C2'] = 50
        
        # Save
        filepath = 'output/test_excel.xlsx'
        wb.save(filepath)
        logger.info(f"✓ Excel file created successfully: {filepath}")
        
        return True
        
    except Exception as e:
        logger.error(f"✗ Excel error: {e}")
        return False


def main():
    """Run all tests."""
    logger.info("="*70)
    logger.info("WEB SCRAPER COMPONENT TESTS")
    logger.info("="*70)
    
    tests = [
        ("Configuration", test_config),
        ("HTML Parsing (BeautifulSoup)", test_beautifulsoup),
        ("Selector Cache", test_selector_cache),
        ("Excel Creation", test_excel_creation),
        ("Data Pipeline", test_data_pipeline),
    ]
    
    results = []
    for name, test_func in tests:
        try:
            result = test_func()
            results.append((name, result))
        except Exception as e:
            logger.error(f"✗ Test {name} crashed: {e}")
            results.append((name, False))
    
    # Summary
    logger.info("\n" + "="*70)
    logger.info("TEST SUMMARY")
    logger.info("="*70)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for name, result in results:
        status = "✓ PASS" if result else "✗ FAIL"
        logger.info(f"{status}: {name}")
    
    logger.info(f"\nTotal: {passed}/{total} tests passed")
    
    if passed == total:
        logger.info("\n✓ All tests passed! System is ready.")
        logger.info("\nNext steps:")
        logger.info("1. Get OpenAI API key from https://platform.openai.com/api-keys")
        logger.info("2. Create .env file with: OPENAI_API_KEY=sk-your-key")
        logger.info("3. Run: python run_scraper.py")
        return 0
    else:
        logger.error(f"\n✗ {total - passed} test(s) failed. Please fix issues.")
        return 1


if __name__ == '__main__':
    sys.exit(main())
